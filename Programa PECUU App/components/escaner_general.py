import socket
import flet as ft
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from services.api_service import ApiService

PORT_LIST = [80, 81, 82, 83, 84, 85, 86, 87, 88, 554, 1024, 1025, 1026, 1027, 1028, 1029]

# Las URLs ahora se cargan dinámicamente desde la API
# Este es un fallback en caso de que la API falle
PREDEFINED_URLS = []

# Variable global para almacenar la ruta del último archivo Excel generado
last_excel_file = None

def scan_ports(ip_addr: str, results_column: ft.Column):
    """
    Intenta establecer una conexión TCP con una dirección IP en la lista de puertos predefinidos.

    Utiliza un socket con un tiempo de espera de 3 segundos para determinar 
    si el puerto está abierto o cerrado.

    :param ip_addr: Dirección IP (v4) a escanear.
    :type ip_addr: str
    :param results_column: Columna de Flet donde se podría mostrar un mensaje de error 
                           de conexión (no de puerto cerrado).
    :type results_column: ft.Column
    :returns: Una tupla que contiene dos listas: puertos abiertos y puertos cerrados.
    :rtype: tuple[list[int], list[int]]
    """
    open_ports = []
    closed_ports = []
    
    for port in PORT_LIST:
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1)
            status = sock.connect_ex((ip_addr, port))
            
            if status == 0:
                open_ports.append(port)
            else:
                closed_ports.append(port)
            sock.close()
            
        except socket.error as err:
            results_column.controls.append(
                ft.Icon(ft.Icons.ERROR, color=ft.Colors.ORANGE),
                ft.Text(f"Puerto {port}: ERROR DE CONEXIÓN", color=ft.Colors.ORANGE, size=16)
            )
            continue
        
    return open_ports, closed_ports


def create_activities_for_no_ports(scan_results: list, results_column: ft.Column, page: ft.Page):
    """
    Crea actividades automáticamente para los enlaces sin puertos abiertos.

    Después de completar el escaneo, esta función identifica los enlaces que no tienen
    ningún puerto abierto y crea una actividad de seguimiento para cada uno.

    :param scan_results: Lista de resultados del escaneo.
    :type scan_results: list
    :param results_column: Columna de Flet donde se mostrarán los resultados finales.
    :type results_column: ft.Column
    :param page: La página principal de Flet, utilizada para forzar las actualizaciones de UI.
    :type page: ft.Page
    """
    # Filtrar los enlaces sin puertos abiertos y con estado exitoso
    no_ports_links = [
        item for item in scan_results 
        if item['status'] == 'success' and not item['open_ports']
    ]
    
    if not no_ports_links:
        print("[INFO] No hay enlaces sin puertos abiertos para crear actividades")
        return
    
    print(f"[INFO] Creando actividades para {len(no_ports_links)} enlaces sin puertos abiertos")
    
    # Crear una actividad por cada enlace sin puertos abiertos
    for link in no_ports_links:
        actividad_data = {
            "titulo": f"Revisar: {link['nombre'] if 'nombre' in link else link['url']}",
            "descripcion": f"El escaneo general detectó que el enlace {link['url']} (IP: {link['ip']}) no tiene ningún puerto abierto en la lista de puertos monitoreados.",
            "fecha": datetime.now().strftime("%Y-%m-%d")
        }
        
        try:
            result = ApiService.create_activity(actividad_data)
            if result["success"]:
                print(f"[SUCCESS] Actividad creada para {link['url']}")
            else:
                print(f"[ERROR] No se pudo crear actividad para {link['url']}: {result['message']}")
        except Exception as ex:
            print(f"[ERROR] Excepción al crear actividad para {link['url']}: {str(ex)}")
    
    print("[INFO] Proceso de creación de actividades completado")


def export_scan_results_to_excel(scan_results: list) -> str:
    """
    Exporta los resultados del escaneo a un archivo Excel.

    Crea un archivo Excel con los resultados del escaneo general, incluyendo
    información detallada sobre cada enlace escaneado, puertos abiertos/cerrados
    y estado del escaneo.

    :param scan_results: Lista de resultados del escaneo.
    :type scan_results: list
    :returns: Ruta del archivo Excel generado.
    :rtype: str
    """
    global last_excel_file
    
    try:
        # Crear workbook y hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Escaneo General"
        
        # Definir estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        title_font = Font(bold=True, size=11)
        success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        error_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados principales
        headers = ["DDNS/URL", "IP", "Estado", "Puertos Abiertos", "Puertos Cerrados", "Total Puertos"]
        ws.append(headers)
        
        # Aplicar estilo a los encabezados
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Agregar datos del escaneo
        row_num = 2
        for item in scan_results:
            if item['status'] == 'success':
                estado = "✓ Exitoso"
                fill = success_fill
                puertos_abiertos = ', '.join(map(str, item['open_ports'])) if item['open_ports'] else "-"
                puertos_cerrados = ', '.join(map(str, item['closed_ports'])) if item['closed_ports'] else "-"
                total_puertos = len(item['closed_ports']) + len(item['open_ports'])
            else:
                estado = f"✗ {item['status']}: {item.get('error', 'Error desconocido')}"
                fill = error_fill
                puertos_abiertos = "-"
                puertos_cerrados = "-"
                total_puertos = 0
            
            row_data = [
                item['url'],
                item.get('ip', '-'),
                estado,
                puertos_abiertos,
                puertos_cerrados,
                total_puertos
            ]
            
            ws.append(row_data)
            
            # Aplicar estilo a la fila
            for cell in ws[row_num]:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row_num += 1
        
        # Ajustar ancho de columnas
        column_widths = [25, 20, 30, 30, 30, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Crear carpeta de reportes si no existe
        reports_folder = os.path.join(os.path.expanduser("~"), "Documents", "SGCC_Reportes")
        if not os.path.exists(reports_folder):
            os.makedirs(reports_folder)
        
        # Generar nombre del archivo con fecha y hora
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(reports_folder, f"escaneo_general_{timestamp}.xlsx")
        
        # Guardar el archivo
        wb.save(file_path)
        
        # Guardar la ruta en la variable global
        last_excel_file = file_path
        
        print(f"[SUCCESS] Archivo Excel generado: {file_path}")
        return file_path
        
    except Exception as ex:
        print(f"[ERROR] Error al generar Excel: {str(ex)}")
        return None


def scan_urls_handler(e, results_column: ft.Column, loading_row: ft.Row, scan_button: ft.FilledButton, page: ft.Page, download_button: ft.FilledButton = None):
    """
    Manejador de eventos para iniciar el Escaneo General de DDNS desde la API.

    Obtiene la lista de DDNS de la API y escanea solo esos enlaces.
    Gestiona la UI (deshabilitación/habilitación del botón, visibilidad de carga) 
    y la lógica de escaneo.

    :param e: Objeto de evento de Flet (el evento de clic del botón).
    :param results_column: Columna de Flet donde se mostrarán los resultados finales.
    :type results_column: ft.Column
    :param loading_row: Fila de Flet que contiene el indicador de carga y el texto de estado.
    :type loading_row: ft.Row
    :param scan_button: El botón de escaneo para ser deshabilitado durante el proceso.
    :type scan_button: ft.FilledButton
    :param page: La página principal de Flet, utilizada para forzar las actualizaciones de UI.
    :type page: ft.Page
    :param download_button: El botón de descarga a habilitar cuando se genere el Excel.
    :type download_button: ft.FilledButton
    """
    # --- Configuración Inicial de UI ---
    scan_button.disabled = True
    scan_button.bgcolor = ft.Colors.GREY_500
    scan_button.color = ft.Colors.GREY_600
    page.update()
    
    results_column.controls.clear()

    loading_indicator = loading_row.controls[0]
    status_text = loading_row.controls[1]

    loading_row.visible = True
    loading_indicator.visible = True
    status_text.value = "Cargando enlaces de la API..."
    status_text.color = ft.Colors.BLACK
    status_text.visible = True
    page.update()

    # Obtener URLs/DDNS desde la API
    print("[INFO] Obteniendo enlaces desde la API...")
    enlaces = ApiService.get_links()
    
    if not enlaces:
        results_column.controls.append(
            ft.Container(
                content=ft.Column([
                    ft.Icon(name=ft.Icons.ERROR, size=50, color=ft.Colors.RED_700),
                    ft.Text("No se pudieron obtener los enlaces desde la API", 
                           weight="bold", color=ft.Colors.RED_700, size=14),
                    ft.Text("Verifica tu conexión a internet", 
                           color=ft.Colors.RED_600, size=12)
                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=20,
                margin=10,
                bgcolor=ft.Colors.RED_50,
                border_radius=10,
                width=500
            )
        )
        loading_row.visible = False
        scan_button.disabled = False
        scan_button.bgcolor = ft.Colors.INDIGO_700
        scan_button.color = ft.Colors.WHITE
        page.update()
        return
    
    # Extraer solo los DDNS
    urls = [enlace.get("ddns", "") for enlace in enlaces if enlace.get("ddns")]
    urls = [url for url in urls if url]  # Filtrar vacíos
    
    print(f"[INFO] Se obtuvieron {len(urls)} DDNS para escanear")
    print(f"[DEBUG] DDNS: {urls}")

    if not urls:
        results_column.controls.append(
            ft.Container(
                content=ft.Column([
                    ft.Text("No hay DDNS configurados en la API para escanear", 
                           weight="bold", color=ft.Colors.ORANGE_700, size=14),
                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=20,
                margin=10,
                bgcolor=ft.Colors.ORANGE_50,
                border_radius=10,
                width=500
            )
        )
        loading_row.visible = False
        scan_button.disabled = False
        scan_button.bgcolor = ft.Colors.INDIGO_700
        scan_button.color = ft.Colors.WHITE
        page.update()
        return

    # Lista para almacenar los resultados del escaneo
    scan_results = []
    total_scanned = 0
    
    status_text.value = "Escaneando DDNS..."
    page.update()
    
    # Escanear cada DDNS
    for url in urls:
        try:
            ip_addr = socket.gethostbyname(url)
            open_ports, closed_ports = scan_ports(ip_addr, results_column)
            total_scanned += 1
            
            scan_results.append({
                'url': url,
                'ip': ip_addr,
                'open_ports': open_ports,
                'closed_ports': closed_ports,
                'status': 'success'
            })
            
            # Actualizar el contador durante el escaneo
            status_text.value = f"Escaneando... {total_scanned}/{len(urls)}"
            page.update()

        except socket.gaierror:
            total_scanned += 1
            # Error al resolver DNS
            scan_results.append({
                'url': url,
                'ip': 'No se pudo resolver',
                'open_ports': [],
                'closed_ports': [],
                'status': 'dns_error',
                'error': 'Error al resolver DNS'
            })
            status_text.value = f"Escaneando... {total_scanned}/{len(urls)}"
            page.update()
            continue
        except Exception as ex:
            total_scanned += 1
            scan_results.append({
                'url': url,
                'ip': 'Error',
                'open_ports': [],
                'closed_ports': [],
                'status': 'error',
                'error': str(ex)
            })
            status_text.value = f"Escaneando... {total_scanned}/{len(urls)}"
            page.update()
            continue

    # Limpiar los resultados anteriores
    results_column.controls.clear()
    
    # Mostrar encabezado con estadísticas
    results_column.controls.append(
        ft.Container(
            content=ft.Column([
                ft.Text("RESULTADOS DEL ESCANEO DE DDNS", 
                       size=20, weight="bold", color=ft.Colors.BLUE_800),
                ft.Text(f"Total de DDNS escaneados: {total_scanned}", 
                       size=16, weight="bold", color=ft.Colors.GREY_600),
                ft.Divider(height=2, color=ft.Colors.GREY_400)
            ]),
            padding=15,
            margin=5,
            bgcolor=ft.Colors.BLUE_50,
            border_radius=10,
            width=500
        )
    )
    
    # Mostrar resultados detallados
    if scan_results:
        for item in scan_results:
            if item['status'] == 'success':
                # DDNS escaneado exitosamente
                container_content = [
                    ft.Row([
                        ft.Icon(name=ft.Icons.PERM_SCAN_WIFI, 
                               color=ft.Colors.GREEN_400 if item['open_ports'] else ft.Colors.RED_400),
                        ft.Text(f"{item['url']}", weight="bold", size=14, color=ft.Colors.GREY_800)
                    ]),
                    ft.Text(f"IP: {item['ip']}", size=12, color=ft.Colors.GREY_500),
                    ft.Text(f"Puertos abiertos: {', '.join(map(str, item['open_ports'])) if item['open_ports'] else 'Ninguno'}", 
                           weight="bold", 
                           color=ft.Colors.GREEN_600 if item['open_ports'] else ft.Colors.GREY_600,
                           size=12),
                    ft.Text(f"Puertos cerrados: {', '.join(map(str, item['closed_ports'])) if item['closed_ports'] else 'N/A'}", 
                           size=12, color=ft.Colors.GREY_600),
                ]
                bg_color = ft.Colors.GREEN_50 if item['open_ports'] else ft.Colors.GREY_100
                border_color = ft.Colors.GREEN_200 if item['open_ports'] else ft.Colors.GREY_200
                
            else:
                # Error en el DDNS
                container_content = [
                    ft.Row([
                        ft.Icon(name=ft.Icons.ERROR, color=ft.Colors.RED_400),
                        ft.Text(f"{item['url']}", weight="bold", size=14, color=ft.Colors.GREY_800)
                    ]),
                    ft.Text(f"Estado: {item['error']}", color=ft.Colors.RED_600, size=12)
                ]
                bg_color = ft.Colors.RED_50
                border_color = ft.Colors.RED_200

            results_column.controls.append(ft.Container(
                content=ft.Column(container_content),
                padding=12,
                margin=3,
                bgcolor=bg_color,
                border=ft.border.all(1, border_color),
                border_radius=8,
                width=500
            ))

    loading_row.visible = False

    # Crear actividades para enlaces sin puertos abiertos
    create_activities_for_no_ports(scan_results, results_column, page)

    # Exportar resultados a Excel
    excel_path = export_scan_results_to_excel(scan_results)

    # Mensaje final
    results_column.controls.append(
        ft.Row(
            [
                ft.Icon(name=ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN_700),
                ft.Text("Escaneo completado.", weight="bold", color=ft.Colors.GREEN_700, size=18)
            ],
            alignment=ft.MainAxisAlignment.CENTER
        )
    )
    
    # Agregar mensaje con la ruta del archivo Excel
    if excel_path:
        # Habilitar el botón de descarga
        if download_button:
            download_button.disabled = False
            download_button.visible = True
        
        results_column.controls.append(
            ft.Container(
                content=ft.Column([
                    ft.Icon(name=ft.Icons.FILE_DOWNLOAD, size=24, color=ft.Colors.GREEN_700),
                    ft.Text("Reporte exportado:", weight="bold", color=ft.Colors.GREEN_700, size=12),
                    ft.Text(excel_path, size=10, color=ft.Colors.BLUE_600)
                ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=12,
                margin=10,
                bgcolor=ft.Colors.GREEN_50,
                border_radius=8,
                width=500
            )
        )
    
    scan_button.disabled = False
    scan_button.bgcolor = ft.Colors.INDIGO_700
    scan_button.color = ft.Colors.WHITE
    page.update()